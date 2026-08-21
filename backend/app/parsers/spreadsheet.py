"""Excel 解析器：以只读模式提取 XLS/XLSX 工作表内容，并限制工作表、行和单元格规模。"""

from datetime import date, datetime
from pathlib import Path
from typing import Any

import xlrd  # type: ignore[import-untyped]
from openpyxl import load_workbook  # type: ignore[import-untyped]

from app.errors import EvidenceValidationError
from app.models.evidence import Evidence
from app.parsers.base import ParseResult
from app.parsers.limits import (
    MAX_SPREADSHEET_CELLS,
    MAX_SPREADSHEET_ROWS,
    MAX_SPREADSHEET_SHEETS,
    validate_zip_container,
)
from app.parsers.text import ensure_text_limit


def _stringify(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value)


class SpreadsheetParser:
    name = "python-excel"
    version = "1.0.0"

    def supports(self, evidence: Evidence) -> bool:
        return evidence.file_extension in {".xls", ".xlsx"}

    def parse(self, source: Path, evidence: Evidence) -> ParseResult:
        if evidence.file_extension == ".xlsx":
            return self._parse_xlsx(source, evidence)
        return self._parse_xls(source, evidence)

    def _parse_xlsx(self, source: Path, evidence: Evidence) -> ParseResult:
        validate_zip_container(source)
        workbook = load_workbook(source, read_only=True, data_only=True, keep_links=False)
        try:
            if len(workbook.worksheets) > MAX_SPREADSHEET_SHEETS:
                raise EvidenceValidationError("EXCEL_SHEET_LIMIT", "Workbook has too many sheets")
            lines: list[str] = []
            sheets: list[dict[str, object]] = []
            total_cells = 0
            for sheet in workbook.worksheets:
                lines.append(f"## Sheet: {sheet.title}")
                row_count = 0
                column_count = 0
                for row_count, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    if row_count > MAX_SPREADSHEET_ROWS:
                        raise EvidenceValidationError(
                            "EXCEL_ROW_LIMIT", "Worksheet has too many rows"
                        )
                    values = [_stringify(value) for value in row]
                    total_cells += len(values)
                    column_count = max(column_count, len(values))
                    if total_cells > MAX_SPREADSHEET_CELLS:
                        raise EvidenceValidationError(
                            "EXCEL_CELL_LIMIT", "Workbook has too many cells"
                        )
                    lines.append("\t".join(values))
                sheets.append(
                    {"name": sheet.title, "row_count": row_count, "column_count": column_count}
                )
            return self._result(evidence, lines, sheets, total_cells)
        finally:
            workbook.close()

    def _parse_xls(self, source: Path, evidence: Evidence) -> ParseResult:
        workbook: Any = xlrd.open_workbook(str(source), on_demand=True)
        try:
            if workbook.nsheets > MAX_SPREADSHEET_SHEETS:
                raise EvidenceValidationError("EXCEL_SHEET_LIMIT", "Workbook has too many sheets")
            lines: list[str] = []
            sheets: list[dict[str, object]] = []
            total_cells = 0
            for sheet in workbook.sheets():
                if sheet.nrows > MAX_SPREADSHEET_ROWS:
                    raise EvidenceValidationError("EXCEL_ROW_LIMIT", "Worksheet has too many rows")
                lines.append(f"## Sheet: {sheet.name}")
                for row_index in range(sheet.nrows):
                    values = [
                        _stringify(sheet.cell_value(row_index, column))
                        for column in range(sheet.ncols)
                    ]
                    total_cells += len(values)
                    if total_cells > MAX_SPREADSHEET_CELLS:
                        raise EvidenceValidationError(
                            "EXCEL_CELL_LIMIT", "Workbook has too many cells"
                        )
                    lines.append("\t".join(values))
                sheets.append(
                    {"name": sheet.name, "row_count": sheet.nrows, "column_count": sheet.ncols}
                )
            return self._result(evidence, lines, sheets, total_cells)
        finally:
            workbook.release_resources()

    def _result(
        self,
        evidence: Evidence,
        lines: list[str],
        sheets: list[dict[str, object]],
        total_cells: int,
    ) -> ParseResult:
        return ParseResult(
            title=evidence.original_filename,
            text=ensure_text_limit("\n".join(lines)),
            language=None,
            metadata={"sheet_count": len(sheets), "sheets": sheets, "cell_count": total_cells},
        )
